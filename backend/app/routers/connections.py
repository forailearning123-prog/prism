import asyncio
import base64
import csv
import io
import time
from datetime import datetime, timedelta, timezone
from typing import Any

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, status
from openpyxl import load_workbook
from sqlalchemy import asc, desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.auth import get_current_user
from app.database import get_db
from app.models import (
    ConnectionLog,
    ConnectionType,
    DataSource,
    DataSourcePermission,
    HealthStatus,
    LogResult,
    MetadataEntry,
    PermissionRole,
    SyncHistory,
    SyncResult,
    Tag,
    User,
)
from app.schemas import (
    ConnectionLogOut,
    ConnectionTestRequest,
    ConnectionTestResponse,
    ConnectorCard,
    DataSourceCreate,
    DataSourceDetailsOut,
    DataSourceListItem,
    DataSourceListResponse,
    DataSourceUpdate,
    HealthStatusEnum,
    MetadataRefreshResponse,
    SourceHealthOut,
    SourceTypeEnum,
    SyncHistoryOut,
)
from app.security import decrypt_sensitive_payload, encrypt_sensitive_payload

router = APIRouter(prefix="/connections", tags=["connections"])
CONNECTOR_CARDS = [
    ConnectorCard(type=SourceTypeEnum.postgresql, title="PostgreSQL", description="Connect to PostgreSQL databases"),
    ConnectorCard(type=SourceTypeEnum.mysql, title="MySQL", description="Connect to MySQL databases"),
    ConnectorCard(type=SourceTypeEnum.sqlserver, title="SQL Server", description="Connect to Microsoft SQL Server"),
    ConnectorCard(type=SourceTypeEnum.csv, title="CSV Upload", description="Upload CSV files"),
    ConnectorCard(type=SourceTypeEnum.excel, title="Excel Upload", description="Upload Excel files"),
    ConnectorCard(type=SourceTypeEnum.rest_api, title="REST API", description="Connect to REST APIs"),
]


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def normalize_error_message(_: Exception) -> str:
    return "Connection failed. Verify details and network access, then try again."


def _build_sensitive_payload(payload: ConnectionTestRequest | DataSourceCreate | DataSourceUpdate) -> dict[str, Any]:
    return {
        "username": getattr(payload, "username", None),
        "password": getattr(payload, "password", None),
        "headers": getattr(payload, "headers", None) or {},
        "file_name": getattr(payload, "file_name", None),
        "file_content_base64": getattr(payload, "file_content_base64", None),
    }


def _to_connection_test_request(source: DataSource, sensitive: dict[str, Any]) -> ConnectionTestRequest:
    return ConnectionTestRequest(
        source_type=SourceTypeEnum(source.source_type.value),
        host=source.host,
        port=source.port,
        username=sensitive.get("username"),
        **{"password": sensitive.get("password")},
        database_name=source.database_name,
        base_url=source.base_url,
        authentication_type=source.authentication_type,
        headers=sensitive.get("headers", {}),
        file_name=sensitive.get("file_name"),
        file_content_base64=sensitive.get("file_content_base64"),
    )


async def _log_action(
    db: AsyncSession,
    source_id: int,
    action: str,
    result: LogResult,
    duration_ms: int,
    user_id: int | None,
    error_message: str = "",
):
    db.add(
        ConnectionLog(
            data_source_id=source_id,
            action=action,
            result=result,
            duration_ms=duration_ms,
            error_message=error_message[:500],
            user_id=user_id,
        )
    )


def _detect_health(source: DataSource) -> tuple[HealthStatus, str]:
    if source.is_deleted:
        return HealthStatus.disconnected, "Connection deleted."

    if not source.last_sync_at:
        return HealthStatus.pending, "Connection has not been tested yet."

    if source.status == HealthStatus.failed:
        return HealthStatus.failed, "Most recent connection check failed."

    if source.last_successful_refresh_at and source.last_successful_refresh_at < utcnow() - timedelta(days=7):
        return HealthStatus.warning, "Metadata refresh is older than 7 days."

    return HealthStatus.healthy, "Connection is healthy."


async def _ensure_access(
    db: AsyncSession,
    source: DataSource,
    user: User,
    action: str,
):
    if user.is_superuser:
        return
    if source.owner_id == user.id:
        return

    permission = await db.scalar(
        select(DataSourcePermission).where(
            DataSourcePermission.data_source_id == source.id,
            DataSourcePermission.user_id == user.id,
        )
    )
    if not permission:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied for this source")

    role = permission.role
    if action == "read":
        return
    if action in {"edit", "test", "refresh"} and role in {PermissionRole.admin, PermissionRole.editor}:
        return
    if action == "delete" and role == PermissionRole.admin:
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")


async def _get_source_or_404(db: AsyncSession, source_id: int) -> DataSource:
    source = await db.scalar(
        select(DataSource)
        .where(DataSource.id == source_id, DataSource.is_deleted.is_(False))
        .options(
            selectinload(DataSource.owner),
            selectinload(DataSource.tags),
            selectinload(DataSource.metadata_entries),
            selectinload(DataSource.sync_history),
            selectinload(DataSource.connection_logs),
        )
    )
    if not source:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data source not found")
    return source


def _decode_base64_content(file_name: str | None, encoded: str | None) -> bytes:
    if not file_name or not encoded:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File upload content is required")
    try:
        return base64.b64decode(encoded, validate=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is not valid base64 content")


def _extract_csv_metadata(file_name: str, encoded: str) -> list[dict[str, Any]]:
    raw = _decode_base64_content(file_name, encoded)
    text = raw.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)[:5]
    metadata = []
    for col in reader.fieldnames or []:
        sample = rows[0].get(col) if rows else None
        metadata.append(
            {
                "object_name": file_name,
                "object_type": "table",
                "column_name": col,
                "data_type": "string",
                "is_nullable": True,
                "sample_value": str(sample)[:500] if sample is not None else None,
            }
        )
    return metadata


def _extract_excel_metadata(file_name: str, encoded: str) -> list[dict[str, Any]]:
    raw = _decode_base64_content(file_name, encoded)
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    headers = []
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row:
        headers = [str(col) for col in first_row if col is not None]
    sample_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    metadata = []
    for idx, col in enumerate(headers):
        sample = sample_row[idx] if sample_row and idx < len(sample_row) else None
        metadata.append(
            {
                "object_name": sheet.title,
                "object_type": "table",
                "column_name": col,
                "data_type": "string",
                "is_nullable": True,
                "sample_value": str(sample)[:500] if sample is not None else None,
            }
        )
    return metadata


async def _extract_rest_metadata(base_url: str, headers: dict[str, str]) -> list[dict[str, Any]]:
    async with httpx.AsyncClient(timeout=10) as client:
        response = await client.get(base_url, headers=headers)
        response.raise_for_status()
        payload = response.json()

    sample = None
    if isinstance(payload, list) and payload:
        sample = payload[0]
    elif isinstance(payload, dict):
        sample = payload
        for value in payload.values():
            if isinstance(value, list) and value:
                sample = value[0]
                break
    if not isinstance(sample, dict):
        return []

    metadata = []
    for key, value in sample.items():
        metadata.append(
            {
                "object_name": "api_payload",
                "object_type": "api_resource",
                "column_name": str(key),
                "data_type": type(value).__name__,
                "is_nullable": value is None,
                "sample_value": str(value)[:500] if value is not None else None,
            }
        )
    return metadata


async def _build_metadata_preview(payload: ConnectionTestRequest) -> list[dict[str, Any]]:
    if payload.source_type == SourceTypeEnum.csv:
        return _extract_csv_metadata(payload.file_name or "", payload.file_content_base64 or "")
    if payload.source_type == SourceTypeEnum.excel:
        return _extract_excel_metadata(payload.file_name or "", payload.file_content_base64 or "")
    if payload.source_type == SourceTypeEnum.rest_api and payload.base_url:
        return await _extract_rest_metadata(payload.base_url, payload.headers)
    return []


async def _test_connection(payload: ConnectionTestRequest) -> ConnectionTestResponse:
    start = time.perf_counter()
    try:
        if payload.source_type in {SourceTypeEnum.postgresql, SourceTypeEnum.mysql, SourceTypeEnum.sqlserver}:
            if not payload.host or not payload.port:
                raise HTTPException(status_code=400, detail="Host and port are required")
            if not payload.username or not payload.password:
                raise HTTPException(status_code=400, detail="Username and password are required")
            _, writer = await asyncio.wait_for(asyncio.open_connection(payload.host, payload.port), timeout=5)
            writer.close()
            await writer.wait_closed()
            message = "Network and authentication inputs validated."
            details = "Connection endpoint reachable. Validate database permissions during metadata refresh."
        elif payload.source_type == SourceTypeEnum.rest_api:
            if not payload.base_url:
                raise HTTPException(status_code=400, detail="Base URL is required")
            async with httpx.AsyncClient(timeout=8) as client:
                response = await client.get(payload.base_url, headers=payload.headers)
                if response.status_code >= 400:
                    raise HTTPException(
                        status_code=400,
                        detail=f"REST API responded with status {response.status_code}. Verify authentication and endpoint.",
                    )
            message = "API endpoint is reachable and returned a valid response."
            details = None
        elif payload.source_type == SourceTypeEnum.csv:
            metadata = _extract_csv_metadata(payload.file_name or "", payload.file_content_base64 or "")
            if not metadata:
                raise HTTPException(status_code=400, detail="CSV file must include headers and at least one column")
            message = "CSV file validated successfully."
            details = None
        elif payload.source_type == SourceTypeEnum.excel:
            metadata = _extract_excel_metadata(payload.file_name or "", payload.file_content_base64 or "")
            if not metadata:
                raise HTTPException(status_code=400, detail="Excel file must include headers and at least one column")
            message = "Excel file validated successfully."
            details = None
        else:
            raise HTTPException(status_code=400, detail="Unsupported connection type")
        duration = int((time.perf_counter() - start) * 1000)
        return ConnectionTestResponse(success=True, message=message, duration_ms=duration, details=details)
    except HTTPException as exc:
        duration = int((time.perf_counter() - start) * 1000)
        return ConnectionTestResponse(success=False, message=exc.detail, duration_ms=duration)
    except Exception as exc:
        duration = int((time.perf_counter() - start) * 1000)
        return ConnectionTestResponse(success=False, message=normalize_error_message(exc), duration_ms=duration)


async def _upsert_tags(db: AsyncSession, source: DataSource, tag_names: list[str]):
    source.tags.clear()
    for raw in tag_names:
        name = raw.strip()
        if not name:
            continue
        tag = await db.scalar(select(Tag).where(func.lower(Tag.name) == name.lower()))
        if not tag:
            tag = Tag(name=name)
            db.add(tag)
            await db.flush()
        source.tags.append(tag)


async def _replace_metadata(db: AsyncSession, source: DataSource, metadata_rows: list[dict[str, Any]]):
    for existing in list(source.metadata_entries):
        await db.delete(existing)
    for row in metadata_rows:
        source.metadata_entries.append(
            MetadataEntry(
                object_name=row["object_name"][:255],
                object_type=row.get("object_type", "table")[:50],
                column_name=row["column_name"][:255],
                data_type=row.get("data_type", "string")[:100],
                is_nullable=bool(row.get("is_nullable", True)),
                sample_value=row.get("sample_value"),
            )
        )


def _to_list_item(source: DataSource) -> DataSourceListItem:
    return DataSourceListItem(
        id=source.id,
        name=source.name,
        source_type=SourceTypeEnum(source.source_type.value),
        status=HealthStatusEnum(source.status.value),
        owner=source.owner.full_name,
        last_sync_at=source.last_sync_at,
        last_successful_refresh_at=source.last_successful_refresh_at,
        created_at=source.created_at,
        tags=[tag.name for tag in source.tags],
    )


def _to_details(source: DataSource) -> DataSourceDetailsOut:
    metadata_rows = [
        {
            "object_name": m.object_name,
            "object_type": m.object_type,
            "column_name": m.column_name,
            "data_type": m.data_type,
            "is_nullable": m.is_nullable,
            "sample_value": m.sample_value,
        }
        for m in sorted(source.metadata_entries, key=lambda row: (row.object_name, row.column_name))
    ]
    recent_syncs = sorted(source.sync_history, key=lambda row: row.created_at, reverse=True)[:10]
    recent_logs = sorted(source.connection_logs, key=lambda row: row.created_at, reverse=True)[:50]

    return DataSourceDetailsOut(
        id=source.id,
        name=source.name,
        source_type=SourceTypeEnum(source.source_type.value),
        status=HealthStatusEnum(source.status.value),
        owner=source.owner.full_name,
        description=source.description,
        schedule=source.schedule,
        host=source.host,
        port=source.port,
        database_name=source.database_name,
        base_url=source.base_url,
        authentication_type=source.authentication_type,
        created_at=source.created_at,
        updated_at=source.updated_at,
        last_sync_at=source.last_sync_at,
        last_successful_refresh_at=source.last_successful_refresh_at,
        tags=[tag.name for tag in source.tags],
        metadata=metadata_rows,
        recent_syncs=[SyncHistoryOut.model_validate(sync) for sync in recent_syncs],
        logs=[ConnectionLogOut.model_validate(log) for log in recent_logs],
        usage_count=source.usage_count,
        recommendations=[],
    )


@router.get("/connectors", response_model=list[ConnectorCard])
async def list_connectors(current_user: User = Depends(get_current_user)):
    return CONNECTOR_CARDS


@router.post("/test", response_model=ConnectionTestResponse)
async def test_connection(payload: ConnectionTestRequest, current_user: User = Depends(get_current_user)):
    return await _test_connection(payload)


@router.post("/preview")
async def preview_connection(payload: ConnectionTestRequest, current_user: User = Depends(get_current_user)):
    test_result = await _test_connection(payload)
    if not test_result.success:
        raise HTTPException(status_code=400, detail=test_result.message)
    preview = await _build_metadata_preview(payload)
    return {"tables": sorted({row["object_name"] for row in preview}), "columns": preview, "sample_records": preview[:5]}


@router.get("/", response_model=DataSourceListResponse)
async def list_connections(
    search: str | None = Query(default=None),
    source_type: SourceTypeEnum | None = Query(default=None),
    status_filter: HealthStatusEnum | None = Query(default=None, alias="status"),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = (
        select(DataSource)
        .where(DataSource.is_deleted.is_(False))
        .options(selectinload(DataSource.owner), selectinload(DataSource.tags))
    )
    if not current_user.is_superuser:
        query = query.outerjoin(
            DataSourcePermission,
            DataSourcePermission.data_source_id == DataSource.id,
        ).where(or_(DataSource.owner_id == current_user.id, DataSourcePermission.user_id == current_user.id))
    if search:
        query = query.where(DataSource.name.ilike(f"%{search.strip()}%"))
    if source_type:
        query = query.where(DataSource.source_type == ConnectionType(source_type.value))
    if status_filter:
        query = query.where(DataSource.status == HealthStatus(status_filter.value))

    order_columns = {
        "name": DataSource.name,
        "created_at": DataSource.created_at,
        "last_sync_at": DataSource.last_sync_at,
        "last_successful_refresh_at": DataSource.last_successful_refresh_at,
        "status": DataSource.status,
    }
    order_col = order_columns.get(sort_by, DataSource.created_at)
    query = query.order_by(asc(order_col) if sort_order == "asc" else desc(order_col))

    count_query = select(func.count()).select_from(query.distinct(DataSource.id).subquery())
    total = int(await db.scalar(count_query) or 0)

    rows = await db.scalars(query.distinct(DataSource.id).offset((page - 1) * page_size).limit(page_size))
    items = [_to_list_item(source) for source in rows.unique().all()]
    return DataSourceListResponse(items=items, total=total, page=page, page_size=page_size)


@router.post("/", response_model=DataSourceDetailsOut, status_code=status.HTTP_201_CREATED)
async def create_connection(
    payload: DataSourceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    duplicate = await db.scalar(
        select(DataSource).where(
            func.lower(DataSource.name) == payload.name.strip().lower(),
            DataSource.is_deleted.is_(False),
        )
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="A data source with this name already exists")

    source = DataSource(
        name=payload.name.strip(),
        source_type=ConnectionType(payload.source_type.value),
        owner_id=current_user.id,
        description=payload.description.strip(),
        schedule=payload.schedule.strip(),
        host=payload.host,
        port=payload.port,
        database_name=payload.database_name,
        base_url=payload.base_url,
        authentication_type=payload.authentication_type,
        settings=payload.settings or {},
        encrypted_credentials=encrypt_sensitive_payload(_build_sensitive_payload(payload)),
    )
    db.add(source)
    await db.flush()

    await _upsert_tags(db, source, payload.tags)
    db.add(DataSourcePermission(data_source_id=source.id, user_id=current_user.id, role=PermissionRole.owner))

    test_payload = ConnectionTestRequest(**payload.model_dump())
    test_result = await _test_connection(test_payload)
    source.status = HealthStatus.healthy if test_result.success else HealthStatus.failed
    source.last_sync_at = utcnow()
    if test_result.success:
        source.last_successful_refresh_at = utcnow()
        preview = await _build_metadata_preview(test_payload)
        await _replace_metadata(db, source, preview)

    db.add(
        SyncHistory(
            data_source_id=source.id,
            result=SyncResult.success if test_result.success else SyncResult.failed,
            duration_ms=test_result.duration_ms,
            message=test_result.message,
            triggered_by=current_user.id,
        )
    )
    await _log_action(
        db=db,
        source_id=source.id,
        action="create_connection",
        result=LogResult.success if test_result.success else LogResult.failed,
        duration_ms=test_result.duration_ms,
        user_id=current_user.id,
        error_message="" if test_result.success else test_result.message,
    )

    await db.commit()
    created = await _get_source_or_404(db, source.id)
    return _to_details(created)


@router.get("/{source_id}", response_model=DataSourceDetailsOut)
async def get_connection_details(
    source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "read")
    source.usage_count += 1
    health, _ = _detect_health(source)
    source.status = health
    await db.commit()

    return _to_details(source)


@router.patch("/{source_id}", response_model=DataSourceDetailsOut)
async def update_connection(
    source_id: int,
    payload: DataSourceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "edit")

    if payload.name and payload.name.strip().lower() != source.name.lower():
        duplicate = await db.scalar(
            select(DataSource).where(
                func.lower(DataSource.name) == payload.name.strip().lower(),
                DataSource.id != source.id,
                DataSource.is_deleted.is_(False),
            )
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="A data source with this name already exists")
        source.name = payload.name.strip()

    for attr in [
        "description",
        "schedule",
        "host",
        "port",
        "database_name",
        "base_url",
        "authentication_type",
    ]:
        value = getattr(payload, attr)
        if value is not None:
            setattr(source, attr, value)

    if payload.settings is not None:
        source.settings = payload.settings

    if payload.tags is not None:
        await _upsert_tags(db, source, payload.tags)

    if any(
        getattr(payload, key) is not None
        for key in ["username", "password", "headers", "file_name", "file_content_base64"]
    ):
        current_sensitive = decrypt_sensitive_payload(source.encrypted_credentials)
        updates = {
            "username": payload.username,
            "password": payload.password,
            "headers": payload.headers,
            "file_name": payload.file_name,
            "file_content_base64": payload.file_content_base64,
        }
        for key, value in updates.items():
            if value is not None:
                current_sensitive[key] = value
        source.encrypted_credentials = encrypt_sensitive_payload(current_sensitive)

    await _log_action(
        db=db,
        source_id=source.id,
        action="update_connection",
        result=LogResult.success,
        duration_ms=0,
        user_id=current_user.id,
    )
    await db.commit()
    updated = await _get_source_or_404(db, source.id)
    return _to_details(updated)


@router.delete("/{source_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_connection(
    source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "delete")
    source.is_deleted = True
    source.deleted_at = utcnow()
    source.status = HealthStatus.disconnected
    await _log_action(
        db=db,
        source_id=source.id,
        action="delete_connection",
        result=LogResult.success,
        duration_ms=0,
        user_id=current_user.id,
    )
    await db.commit()
    return None


@router.post("/{source_id}/test", response_model=ConnectionTestResponse)
async def test_saved_connection(
    source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "test")
    sensitive = decrypt_sensitive_payload(source.encrypted_credentials)
    payload = _to_connection_test_request(source, sensitive)
    test_result = await _test_connection(payload)

    source.last_sync_at = utcnow()
    source.status = HealthStatus.healthy if test_result.success else HealthStatus.failed
    if test_result.success:
        source.last_successful_refresh_at = utcnow()

    db.add(
        SyncHistory(
            data_source_id=source.id,
            result=SyncResult.success if test_result.success else SyncResult.failed,
            duration_ms=test_result.duration_ms,
            message=test_result.message,
            triggered_by=current_user.id,
        )
    )
    await _log_action(
        db=db,
        source_id=source.id,
        action="test_connection",
        result=LogResult.success if test_result.success else LogResult.failed,
        duration_ms=test_result.duration_ms,
        user_id=current_user.id,
        error_message="" if test_result.success else test_result.message,
    )
    await db.commit()
    return test_result


@router.post("/{source_id}/refresh-metadata", response_model=MetadataRefreshResponse)
async def refresh_metadata(
    source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "refresh")
    sensitive = decrypt_sensitive_payload(source.encrypted_credentials)
    payload = _to_connection_test_request(source, sensitive)

    test_result = await _test_connection(payload)
    if not test_result.success:
        source.status = HealthStatus.failed
        await _log_action(
            db=db,
            source_id=source.id,
            action="refresh_metadata",
            result=LogResult.failed,
            duration_ms=test_result.duration_ms,
            user_id=current_user.id,
            error_message=test_result.message,
        )
        await db.commit()
        raise HTTPException(status_code=400, detail=test_result.message)

    metadata_rows = await _build_metadata_preview(payload)
    await _replace_metadata(db, source, metadata_rows)

    source.last_sync_at = utcnow()
    source.last_successful_refresh_at = utcnow()
    source.status = HealthStatus.healthy
    db.add(
        SyncHistory(
            data_source_id=source.id,
            result=SyncResult.success,
            duration_ms=test_result.duration_ms,
            message="Metadata refreshed successfully",
            triggered_by=current_user.id,
        )
    )
    await _log_action(
        db=db,
        source_id=source.id,
        action="refresh_metadata",
        result=LogResult.success,
        duration_ms=test_result.duration_ms,
        user_id=current_user.id,
    )
    await db.commit()
    return MetadataRefreshResponse(
        source_id=source.id,
        status="success",
        metadata_items=len(metadata_rows),
        message="Metadata refreshed successfully",
    )


@router.get("/{source_id}/logs", response_model=list[ConnectionLogOut])
async def get_connection_logs(
    source_id: int,
    limit: int = Query(default=100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "read")
    rows = await db.scalars(
        select(ConnectionLog)
        .where(ConnectionLog.data_source_id == source_id)
        .order_by(ConnectionLog.created_at.desc())
        .limit(limit)
    )
    return [ConnectionLogOut.model_validate(row) for row in rows.all()]


@router.get("/{source_id}/health", response_model=SourceHealthOut)
async def get_connection_health(
    source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    source = await _get_source_or_404(db, source_id)
    await _ensure_access(db, source, current_user, "read")
    status_value, reason = _detect_health(source)
    source.status = status_value
    await db.commit()
    return SourceHealthOut(source_id=source.id, status=HealthStatusEnum(status_value.value), reason=reason)
